from openpyxl import load_workbook


def read_zones(filename):

    workbook = load_workbook(filename=filename)
    sheet = workbook["Sheet1"]

    zones_list = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None:
                value = cell.value.split(";")
                zones_list.extend(value)
       
    return zones_list
                
                


def read_sec_rules(filename):
    workbook = load_workbook(filename=filename)
    sheet = workbook["Security Rules"]

    data = {}
    
    for col in sheet.iter_cols(min_col=2, max_col=sheet.max_column, min_row=0, max_row=sheet.max_row):
        key = col[0].value
        zones_list =[]

        for cell in col[1:]:
            if cell.value is not None:
                value = cell.value
                zones_list.append(value)
                data[key] = zones_list

    #returns data in dictinary file. Column names are keys, and each column cells are values
    return data


